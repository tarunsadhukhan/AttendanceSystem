"""
Other Entries module endpoints (Production → Finishing Entry → Other Entries).
Table: tbl_daily_finishing
Columns: tran_date, spell_id, looms, cuts, hemming, heracle, cutting, branding,
         hand_sewer, bales, issue_bales, updated_by, updated_date_time
"""
from datetime import datetime

from flask import request, jsonify, send_file
from . import otherentries_bp
from src.database import get_db
from src.send_whatsapp import send_text


_MIGRATED = False

# Field -> WhatsApp label, in display order (same items as the SPG report).
_NOTIFY_FIELDS = [
    ('looms', 'Looms'), ('cuts', 'Cuts'), ('hemming', 'Hemming'),
    ('heracle', 'Heracle'), ('cutting', 'Cutting'), ('branding', 'Branding'),
    ('hand_sewer', 'Hand Sewer'), ('bales', 'Bales'), ('issue_bales', 'Issue Bales'),
]


def _wa_clean_number(mobno):
    """Digits only; bare 10-digit numbers get the default country code."""
    import os
    if not mobno:
        return None
    digits = ''.join(ch for ch in str(mobno) if ch.isdigit())
    if not digits:
        return None
    if len(digits) == 10:
        digits = os.environ.get('WHATSAPP_DEFAULT_CC', '91') + digits
    return digits


def _notify_other_entry(action, tran_date, spell_id, values):
    """WhatsApp the msg_for='OE' recipients that an Other Entries record was
    saved/updated. Best-effort: any failure is logged, never raised."""
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT name, mobno FROM tbl_whatsapp_send WHERE msg_for = 'OE'")
        recipients = cur.fetchall()
        spell_name = ''
        if spell_id:
            cur.execute("SELECT spell_name FROM spell_mst WHERE spell_id = %s", (spell_id,))
            row = cur.fetchone()
            spell_name = (row or {}).get('spell_name') or ''
        cur.close(); db.close()
    except Exception as ex:
        print('OtherEntries WhatsApp lookup failed:', ex)
        return
    if not recipients:
        print('OtherEntries WhatsApp: no recipients (tbl_whatsapp_send msg_for=OE)')
        return

    date_txt = str(tran_date or '')
    try:
        date_txt = datetime.strptime(date_txt[:10], '%Y-%m-%d').strftime('%d-%m-%Y')
    except ValueError:
        pass
    parts = ['Other Entries %s : Date %s' % (action, date_txt)]
    if spell_name:
        parts.append('Spell %s' % spell_name)
    for col, label in _NOTIFY_FIELDS:
        if values.get(col) is not None:
            parts.append('%s : %s' % (label, values[col]))
    parts.append('At Time : %s' % datetime.now().strftime('%d-%m-%Y %H:%M'))
    body = ' , '.join(parts)

    for r in recipients:
        to_number = _wa_clean_number(r.get('mobno'))
        if not to_number:
            print('OtherEntries WhatsApp: invalid mobno:', r.get('mobno'))
            continue
        ok, info = send_text(to_number, body)
        print('OtherEntries WhatsApp: send to', to_number, '->',
              'OK' if ok else 'FAILED', '|', info)


def _ensure_columns():
    """Add cutting / branding / issue_bales columns if they don't exist yet."""
    global _MIGRATED
    if _MIGRATED:
        return
    try:
        db = get_db()
        cursor = db.cursor()
        for col in ('cutting', 'branding', 'issue_bales'):
            try:
                cursor.execute(
                    f"ALTER TABLE tbl_daily_finishing ADD COLUMN {col} INT DEFAULT NULL"
                )
                db.commit()
            except Exception:
                # column already exists (or other benign error) — ignore
                pass
        cursor.close()
        db.close()
        _MIGRATED = True
    except Exception as e:
        print(f"[otherentries] migration check failed (non-fatal): {e}")


def _to_int(v):
    if v is None or v == '':
        return None
    try:
        return int(float(v))
    except Exception:
        return None


@otherentries_bp.route('/entry', methods=['POST'])
def save_other_entry():
    try:
        _ensure_columns()
        data = request.get_json(silent=True) or request.form.to_dict() or {}
        tran_date = data.get('date')
        spell_id  = data.get('spell_id')
        user_id   = data.get('user_id')

        if not tran_date or not spell_id:
            return jsonify({'status': 'error',
                            'message': 'date and spell_id are required',
                            'received': data}), 400

        looms       = _to_int(data.get('looms'))
        cuts        = _to_int(data.get('cuts'))
        hemming     = _to_int(data.get('hemming'))
        heracle     = _to_int(data.get('heracle'))
        cutting     = _to_int(data.get('cutting'))
        branding    = _to_int(data.get('branding'))
        hand_sewer  = _to_int(data.get('hand_sewer'))
        bales       = _to_int(data.get('bales'))
        issue_bales = _to_int(data.get('issue_bales'))

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            SELECT tbl_daily_fng_id FROM tbl_daily_finishing
            WHERE tran_date = %s AND spell_id = %s LIMIT 1
        """, (tran_date, spell_id))
        if cursor.fetchone():
            cursor.close()
            db.close()
            return jsonify({'status': 'error',
                            'message': 'Entry already exists for this date and spell!'}), 409
        cursor.execute("""
            INSERT INTO tbl_daily_finishing
                (tran_date, spell_id, looms, cuts, hemming, heracle,
                 cutting, branding, hand_sewer, bales, issue_bales, updated_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (tran_date, spell_id, looms, cuts, hemming, heracle,
              cutting, branding, hand_sewer, bales, issue_bales, user_id))
        db.commit()
        new_id = cursor.lastrowid
        cursor.close()
        db.close()
        _notify_other_entry('Saved', tran_date, spell_id, {
            'looms': looms, 'cuts': cuts, 'hemming': hemming, 'heracle': heracle,
            'cutting': cutting, 'branding': branding, 'hand_sewer': hand_sewer,
            'bales': bales, 'issue_bales': issue_bales,
        })
        return jsonify({'status': 'success', 'message': 'Saved', 'id': new_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/entry', methods=['GET'])
def list_other_entries():
    try:
        _ensure_columns()
        date_str = request.args.get('date')
        spell_id = request.args.get('spell_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = """
            SELECT f.tbl_daily_fng_id  AS id,
                   f.tran_date         AS entry_date,
                   f.spell_id,
                   COALESCE(s.spell_name, '') AS spell_name,
                   f.looms,
                   f.cuts,
                   f.hemming,
                   f.heracle,
                   f.cutting,
                   f.branding,
                   f.hand_sewer,
                   f.bales,
                   f.issue_bales
            FROM tbl_daily_finishing f
            LEFT JOIN spell_mst s ON f.spell_id = s.spell_id
            WHERE f.tran_date = %s
        """
        params = [date_str]
        if spell_id:
            query += " AND f.spell_id = %s"
            params.append(spell_id)
        query += " ORDER BY f.tbl_daily_fng_id DESC"
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        for r in rows:
            if r.get('entry_date') is not None:
                r['entry_date'] = r['entry_date'].isoformat()
        return jsonify({'status': 'success', 'entries': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/entry/<int:entry_id>', methods=['PUT'])
def update_other_entry(entry_id):
    try:
        _ensure_columns()
        data = request.get_json(silent=True) or {}
        sets = []
        params = []
        for src_key, col in [
            ('spell_id',    'spell_id'),
            ('looms',       'looms'),
            ('cuts',        'cuts'),
            ('hemming',     'hemming'),
            ('heracle',     'heracle'),
            ('cutting',     'cutting'),
            ('branding',    'branding'),
            ('hand_sewer',  'hand_sewer'),
            ('bales',       'bales'),
            ('issue_bales', 'issue_bales'),
            ('user_id',     'updated_by'),
        ]:
            if src_key in data:
                sets.append(f"{col} = %s")
                params.append(_to_int(data.get(src_key)))
        if 'date' in data:
            sets.append("tran_date = %s")
            params.append(data.get('date'))
        if not sets:
            return jsonify({'status': 'error', 'message': 'no fields to update'}), 400
        sets.append("updated_date_time = CURRENT_TIMESTAMP")
        params.append(entry_id)

        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute(
            "SELECT tran_date, spell_id FROM tbl_daily_finishing WHERE tbl_daily_fng_id = %s",
            (entry_id,))
        current = cursor.fetchone()
        if not current:
            cursor.close()
            db.close()
            return jsonify({'status': 'error', 'message': 'entry not found'}), 404
        new_date  = data.get('date') if 'date' in data else current['tran_date']
        new_spell = _to_int(data.get('spell_id')) if 'spell_id' in data else current['spell_id']
        cursor.execute("""
            SELECT tbl_daily_fng_id FROM tbl_daily_finishing
            WHERE tran_date = %s AND spell_id = %s AND tbl_daily_fng_id != %s LIMIT 1
        """, (new_date, new_spell, entry_id))
        if cursor.fetchone():
            cursor.close()
            db.close()
            return jsonify({'status': 'error',
                            'message': 'Entry already exists for this date and spell!'}), 409
        cursor.close()
        cursor = db.cursor()
        cursor.execute(
            f"UPDATE tbl_daily_finishing SET {', '.join(sets)} WHERE tbl_daily_fng_id = %s",
            tuple(params)
        )
        db.commit()
        rows = cursor.rowcount
        cursor.close()
        db.close()
        if rows == 0:
            return jsonify({'status': 'error', 'message': 'entry not found'}), 404

        # Notify with the row's current (post-update) values.
        try:
            db = get_db()
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM tbl_daily_finishing WHERE tbl_daily_fng_id = %s",
                (entry_id,))
            row = cursor.fetchone() or {}
            cursor.close()
            db.close()
            _notify_other_entry('Updated', row.get('tran_date'),
                                row.get('spell_id'), row)
        except Exception as ex:
            print('OtherEntries WhatsApp notify failed:', ex)

        return jsonify({'status': 'success', 'message': 'Updated', 'id': entry_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/entry/<int:entry_id>', methods=['DELETE'])
def delete_other_entry(entry_id):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "DELETE FROM tbl_daily_finishing WHERE tbl_daily_fng_id = %s",
            (entry_id,)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'message': 'Deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# Daily Machine Summary  (table: tbl_daily_summ_mechine_data)
#   inputs : tran_date, mc_code_id, branch_id, spell_a1/a2/b1/b2/c
#   derived: shift_a = a1 + a2, shift_b = b1 + b2, shift_c = c
#   mc_code_id -> mechine_code_master.mc_code_id
# ─────────────────────────────────────────────────────────────────────────────

def _to_dec(v):
    """Parse a decimal value; '' / None -> None."""
    if v is None or v == '':
        return None
    try:
        return float(v)
    except Exception:
        return None


def _shift_sum(*vals):
    """Sum the given spell values, treating None as 0.
    Returns None only when every value is None (nothing entered)."""
    present = [v for v in vals if v is not None]
    if not present:
        return None
    return round(sum(present), 2)


@otherentries_bp.route('/machine-summary', methods=['POST'])
def save_machine_summary():
    """Insert a daily machine-summary row.
    Body: {date, branch_id, mc_code_id, spell_a1, spell_a2, spell_b1,
           spell_b2, spell_c, user_id}
    """
    try:
        data       = request.get_json(silent=True) or request.form.to_dict() or {}
        tran_date  = data.get('date')
        branch_id  = data.get('branch_id')
        mc_code_id = data.get('mc_code_id')
        if not tran_date or not mc_code_id:
            return jsonify({'status': 'error',
                            'message': 'date and mc_code_id are required'}), 400

        a1 = _to_dec(data.get('spell_a1'))
        a2 = _to_dec(data.get('spell_a2'))
        b1 = _to_dec(data.get('spell_b1'))
        b2 = _to_dec(data.get('spell_b2'))
        c  = _to_dec(data.get('spell_c'))
        shift_a = _shift_sum(a1, a2)
        shift_b = _shift_sum(b1, b2)
        shift_c = _shift_sum(c)

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO tbl_daily_summ_mechine_data
                (tran_date, branch_id, mc_code_id,
                 spell_a1, spell_a2, spell_b1, spell_b2, spell_c,
                 shift_a, shift_b, shift_c, is_active, created_on)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, CURRENT_TIMESTAMP)
        """, (tran_date, branch_id, mc_code_id,
              a1, a2, b1, b2, c, shift_a, shift_b, shift_c))
        db.commit()
        new_id = cursor.lastrowid
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'message': 'Saved', 'id': new_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/machine-summary', methods=['GET'])
def list_machine_summary():
    """List machine-summary rows for a tran_date (+ optional branch_id)."""
    try:
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = """
            SELECT d.daily_sum_mc_id AS id,
                   d.tran_date,
                   d.mc_code_id,
                   m.mc_code,
                   m.Mechine_type_name AS mechine_type_name,
                   d.spell_a1, d.spell_a2, d.spell_b1, d.spell_b2, d.spell_c,
                   d.shift_a, d.shift_b, d.shift_c
            FROM tbl_daily_summ_mechine_data d
            LEFT JOIN mechine_code_master m ON m.mc_code_id = d.mc_code_id
            WHERE d.tran_date = %s
              AND (d.is_active IS NULL OR d.is_active = 1)
        """
        params = [date_str]
        if branch_id:
            query += " AND d.branch_id = %s"
            params.append(branch_id)
        query += " ORDER BY d.daily_sum_mc_id DESC"
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        for r in rows:
            if r.get('tran_date') is not None:
                r['tran_date'] = r['tran_date'].isoformat()
        return jsonify({'status': 'success', 'entries': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/machine-departments', methods=['GET'])
def list_machine_departments():
    """Departments (dept_mst) that have machines, for the New Machine Entry grid.

    mechine_code_master.dept_id references dept_mst.dept_id, so the department
    dropdown must come from dept_mst (not the sub-department list). Only depts
    that actually have active machines are returned.
    Query params: ?branch_id=<id> (required)
    Returns: { status, total, data: [ { id, name } ] }
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        if not branch_id:
            return jsonify({'status': 'error', 'message': 'branch_id is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT DISTINCT d.dept_id AS id, d.dept_desc AS name
            FROM dept_mst d
            JOIN mechine_code_master m ON m.dept_id = d.dept_id
            WHERE d.branch_id = %s
              AND m.branch_id = %s
              AND (m.is_active IS NULL OR m.is_active = 1)
            ORDER BY d.dept_desc
        """, (branch_id, branch_id))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'total': len(rows), 'data': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/hands-report', methods=['GET'])
def hands_report():
    """Hands Report for a date + branch, from vw_hands_report.

    Matches the SJM paper form / ERP "Daily Man-Machine" export: one row per
    designation (particular), pivoted Shift {A, B1, B2, C} x Shed {Old, New}.
    Per cell: M/H = active machine count (mh_*), Hands = SUM(working_hours)/8
    (hands_*). total_hands is the honest head-count (B not double-counted).
    Query params: ?date=YYYY-MM-DD (required) & branch_id=<id> (optional)
    Returns: { status, total, rows: [ { dept_desc, dept_code, particular,
               mh_a_os, hands_a_os, ... mh_c_ns, hands_c_ns, total_hands } ] }
    """
    try:
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = """
            SELECT dept_desc, dept_code, particular,
                   mh_a_os,  hands_a_os,  mh_a_ns,  hands_a_ns,
                   mh_b1_os, hands_b1_os, mh_b1_ns, hands_b1_ns,
                   mh_b2_os, hands_b2_os, mh_b2_ns, hands_b2_ns,
                   mh_c_os,  hands_c_os,  mh_c_ns,  hands_c_ns,
                   total_hands
            FROM vw_hands_report
            WHERE tran_date = %s
        """
        params = [date_str]
        if branch_id:
            query += " AND branch_id = %s"
            params.append(branch_id)
        query += (" ORDER BY CAST(dept_code AS UNSIGNED), "
                  "desig_code IS NULL, CAST(desig_code AS UNSIGNED), particular")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()

        # vw_hands_report returns Decimals; coerce to float for JSON.
        measures = ('mh_a_os', 'hands_a_os', 'mh_a_ns', 'hands_a_ns',
                    'mh_b1_os', 'hands_b1_os', 'mh_b1_ns', 'hands_b1_ns',
                    'mh_b2_os', 'hands_b2_os', 'mh_b2_ns', 'hands_b2_ns',
                    'mh_c_os', 'hands_c_os', 'mh_c_ns', 'hands_c_ns', 'total_hands')
        for r in rows:
            for k in measures:
                r[k] = None if r.get(k) is None else float(r[k])

        return jsonify({'status': 'success', 'total': len(rows), 'rows': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# Measures pivoted per (shift, shed) — display order, cols B..Q of the sheet.
_HANDS_MEAS = ('mh_a_os', 'hands_a_os', 'mh_a_ns', 'hands_a_ns',
               'mh_b1_os', 'hands_b1_os', 'mh_b1_ns', 'hands_b1_ns',
               'mh_b2_os', 'hands_b2_os', 'mh_b2_ns', 'hands_b2_ns',
               'mh_c_os', 'hands_c_os', 'mh_c_ns', 'hands_c_ns')


def _build_hands_report_xlsx(rows, date_str):
    """Build the SJM Hands Report workbook (matches the ERP export) -> bytes.

    rows: vw_hands_report dicts ordered by dept/particular, each with dept_desc,
    particular, the 16 _HANDS_MEAS keys and total_hands.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def z(v):
        v = float(v or 0)
        return None if v == 0 else (int(v) if v == int(v) else round(v, 2))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hands Report"

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1565C0')
    hdr_font = Font(bold=True, color='FFFFFF')
    dept_fill = PatternFill('solid', fgColor='BBDEFB')
    dtot_fill = PatternFill('solid', fgColor='E3F2FD')
    gtot_fill = PatternFill('solid', fgColor='1565C0')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.cell(1, 1, f"HANDS REPORT - Date: {date_str}").font = Font(bold=True)
    ws.merge_cells('A1:Q1')
    ws.cell(1, 1).alignment = center

    # 3-row grouped header (rows 2..4): Shift band -> Old/New Shed -> M/H·Hands.
    ws.merge_cells('A2:A4')
    ws.cell(2, 1, "Particular's")
    for i, s in enumerate(("Shift A", "Shift B1", "Shift B2", "Shift C")):
        c0 = 2 + i * 4                                   # B, F, J, N
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + 3)
        ws.cell(2, c0, s)
        for j, shed in enumerate(("Old Shed", "New Shed")):
            cs = c0 + j * 2
            ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs + 1)
            ws.cell(3, cs, shed)
        for j, m in enumerate(("M/H", "Hands") * 2):
            ws.cell(4, c0 + j, m)
    for r in (2, 3, 4):
        for c in range(1, 18):
            cell = ws.cell(r, c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border

    rownum = 5

    def write_row(label, values, fill=None, bold=False):
        nonlocal rownum
        a = ws.cell(rownum, 1, label)
        a.alignment = left; a.border = border
        if bold:
            a.font = Font(bold=True)
        if fill:
            a.fill = fill
        for i, v in enumerate(values):
            cell = ws.cell(rownum, 2 + i, v)
            cell.alignment = center; cell.border = border
            if bold:
                cell.font = Font(bold=True, color='FFFFFF' if fill is gtot_fill else '000000')
            if fill:
                cell.fill = fill
        rownum += 1

    grand = [0.0] * 16
    gth = 0.0
    i = 0
    while i < len(rows):
        dept = rows[i].get('dept_desc') or '-'
        for c in range(1, 18):                           # dept section header band
            ws.cell(rownum, c).fill = dept_fill; ws.cell(rownum, c).border = border
        hc = ws.cell(rownum, 1, dept)
        hc.font = Font(bold=True); hc.alignment = left
        rownum += 1
        dsum = [0.0] * 16
        while i < len(rows) and (rows[i].get('dept_desc') or '-') == dept:
            r = rows[i]
            write_row(r.get('particular') or '', [z(r.get(k)) for k in _HANDS_MEAS])
            for k, key in enumerate(_HANDS_MEAS):
                dsum[k] += float(r.get(key) or 0)
            gth += float(r.get('total_hands') or 0)
            i += 1
        for k in range(16):
            grand[k] += dsum[k]
        write_row(f"{dept} Total", [z(v) for v in dsum], fill=dtot_fill, bold=True)

    write_row("Grand Total", [z(v) for v in grand], fill=gtot_fill, bold=True)
    hands_sum = sum(grand[k] for k in range(1, 16, 2))
    write_row("Grand Total Hands (as per data)", [z(gth)] + [None] * 15,
              fill=dtot_fill, bold=True)
    write_row("Difference (B counted in B1 & B2)", [z(hands_sum - gth)] + [None] * 15,
              fill=dtot_fill, bold=True)

    ws.column_dimensions['A'].width = 24
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.freeze_panes = 'B5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@otherentries_bp.route('/hands-report.xlsx', methods=['GET'])
def hands_report_xlsx():
    """Hands Report as a downloadable .xlsx (same layout as the ERP export).
    Query params: ?date=YYYY-MM-DD (required) & branch_id=<id> (optional)."""
    try:
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = ("SELECT dept_desc, dept_code, particular, " + ", ".join(_HANDS_MEAS) +
                 ", total_hands FROM vw_hands_report WHERE tran_date = %s")
        params = [date_str]
        if branch_id:
            query += " AND branch_id = %s"
            params.append(branch_id)
        query += (" ORDER BY CAST(dept_code AS UNSIGNED), "
                  "desig_code IS NULL, CAST(desig_code AS UNSIGNED), particular")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()

        buf = _build_hands_report_xlsx(rows, date_str)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"HandsReport_{date_str}.xlsx",
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# Actual-vs-Std base measures per shift (from vw_hands_std_report).
_HANDS_STD_MEAS = ('act_a', 'std_a', 'act_b', 'std_b', 'act_c', 'std_c')


def _build_hands_std_xlsx(rows, date_str):
    """Build the Actual-vs-Std Hands workbook (matches the ERP export) -> bytes.

    Per designation, per Shift {A, B, C}: Act hands, Std hands, Excess/Short
    (= Act - Std), plus a Total group across the three shifts.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def z(v):
        v = round(float(v or 0), 2)
        return None if v == 0 else (int(v) if v == int(v) else v)

    def cells(a_a, s_a, a_b, s_b, a_c, s_c):        # -> 12 display cells (B..M)
        at, st = a_a + a_b + a_c, s_a + s_b + s_c
        return [a_a, s_a, a_a - s_a, a_b, s_b, a_b - s_b,
                a_c, s_c, a_c - s_c, at, st, at - st]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actual vs Std"

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1565C0')
    hdr_font = Font(bold=True, color='FFFFFF')
    dept_fill = PatternFill('solid', fgColor='BBDEFB')
    dtot_fill = PatternFill('solid', fgColor='E3F2FD')
    gtot_fill = PatternFill('solid', fgColor='1565C0')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.cell(1, 1, f"ACTUAL vs STD HANDS - Date: {date_str}").font = Font(bold=True)
    ws.merge_cells('A1:M1')
    ws.cell(1, 1).alignment = center

    # 2-row grouped header (rows 2..3): Shift/Total band -> Act·Std·Excess.
    ws.merge_cells('A2:A3')
    ws.cell(2, 1, "Particular's")
    for i, s in enumerate(("Shift A", "Shift B", "Shift C", "Total")):
        c0 = 2 + i * 3                               # B, E, H, K
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + 2)
        ws.cell(2, c0, s)
        for j, m in enumerate(("Act hands", "Std hands", "Excess/Short")):
            ws.cell(3, c0 + j, m)
    for r in (2, 3):
        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border

    rownum = 4

    def write_row(label, values, fill=None, bold=False):
        nonlocal rownum
        a = ws.cell(rownum, 1, label)
        a.alignment = left; a.border = border
        if bold:
            a.font = Font(bold=True)
        if fill:
            a.fill = fill
        for i, v in enumerate(values):
            cell = ws.cell(rownum, 2 + i, v)
            cell.alignment = center; cell.border = border
            if bold:
                cell.font = Font(bold=True, color='FFFFFF' if fill is gtot_fill else '000000')
            if fill:
                cell.fill = fill
        rownum += 1

    grand = [0.0] * 6
    i = 0
    while i < len(rows):
        dept = rows[i].get('dept_desc') or '-'
        for c in range(1, 14):
            ws.cell(rownum, c).fill = dept_fill; ws.cell(rownum, c).border = border
        hc = ws.cell(rownum, 1, dept)
        hc.font = Font(bold=True); hc.alignment = left
        rownum += 1
        dsum = [0.0] * 6
        while i < len(rows) and (rows[i].get('dept_desc') or '-') == dept:
            r = rows[i]
            base = [float(r.get(k) or 0) for k in _HANDS_STD_MEAS]
            for k in range(6):
                dsum[k] += base[k]
            write_row(r.get('particular') or '', [z(v) for v in cells(*base)])
            i += 1
        for k in range(6):
            grand[k] += dsum[k]
        write_row(f"{dept} Total", [z(v) for v in cells(*dsum)], fill=dtot_fill, bold=True)

    write_row("Grand Total", [z(v) for v in cells(*grand)], fill=gtot_fill, bold=True)

    ws.column_dimensions['A'].width = 24
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@otherentries_bp.route('/hands-std-report.xlsx', methods=['GET'])
def hands_std_report_xlsx():
    """Actual-vs-Std Hands report as a downloadable .xlsx (same as the ERP export).
    Query params: ?date=YYYY-MM-DD (required) & branch_id=<id> (optional)."""
    try:
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = ("SELECT dept_desc, dept_code, particular, " + ", ".join(_HANDS_STD_MEAS) +
                 " FROM vw_hands_std_report WHERE tran_date = %s")
        params = [date_str]
        if branch_id:
            query += " AND branch_id = %s"
            params.append(branch_id)
        query += (" ORDER BY CAST(dept_code AS UNSIGNED), "
                  "desig_code IS NULL, CAST(desig_code AS UNSIGNED), particular")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()

        buf = _build_hands_std_xlsx(rows, date_str)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"ActualVsStdHands_{date_str}.xlsx",
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/hands-std-report', methods=['GET'])
def hands_std_report():
    """Actual-vs-Std Hands report JSON for a date + branch, from vw_hands_std_report.

    One row per designation (particular) with per-shift act_/std_ hands; the FE
    derives Excess/Short (= act - std) and the Total group.
    Query params: ?date=YYYY-MM-DD (required) & branch_id=<id> (optional).
    """
    try:
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not date_str:
            return jsonify({'status': 'error', 'message': 'date is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        query = ("SELECT dept_desc, dept_code, particular, " + ", ".join(_HANDS_STD_MEAS) +
                 " FROM vw_hands_std_report WHERE tran_date = %s")
        params = [date_str]
        if branch_id:
            query += " AND branch_id = %s"
            params.append(branch_id)
        query += (" ORDER BY CAST(dept_code AS UNSIGNED), "
                  "desig_code IS NULL, CAST(desig_code AS UNSIGNED), particular")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()
        db.close()

        for r in rows:
            for k in _HANDS_STD_MEAS:
                r[k] = None if r.get(k) is None else float(r[k])

        return jsonify({'status': 'success', 'total': len(rows), 'rows': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/machine-summary/<int:entry_id>', methods=['PUT'])
def update_machine_summary(entry_id):
    """Update a machine-summary row; recompute shift_* from the spell values."""
    try:
        data       = request.get_json(silent=True) or {}
        a1 = _to_dec(data.get('spell_a1'))
        a2 = _to_dec(data.get('spell_a2'))
        b1 = _to_dec(data.get('spell_b1'))
        b2 = _to_dec(data.get('spell_b2'))
        c  = _to_dec(data.get('spell_c'))
        shift_a = _shift_sum(a1, a2)
        shift_b = _shift_sum(b1, b2)
        shift_c = _shift_sum(c)
        mc_code_id = data.get('mc_code_id')

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            UPDATE tbl_daily_summ_mechine_data
               SET mc_code_id = %s,
                   spell_a1 = %s, spell_a2 = %s,
                   spell_b1 = %s, spell_b2 = %s, spell_c = %s,
                   shift_a = %s, shift_b = %s, shift_c = %s,
                   updated = 'Y'
             WHERE daily_sum_mc_id = %s
        """, (mc_code_id, a1, a2, b1, b2, c, shift_a, shift_b, shift_c, entry_id))
        db.commit()
        rows = cursor.rowcount
        cursor.close()
        db.close()
        if rows == 0:
            return jsonify({'status': 'error', 'message': 'entry not found'}), 404
        return jsonify({'status': 'success', 'message': 'Updated', 'id': entry_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@otherentries_bp.route('/machine-summary/<int:entry_id>', methods=['DELETE'])
def delete_machine_summary(entry_id):
    """Soft-delete (is_active = 0) a machine-summary row."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "UPDATE tbl_daily_summ_mechine_data SET is_active = 0 WHERE daily_sum_mc_id = %s",
            (entry_id,)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'message': 'Deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
