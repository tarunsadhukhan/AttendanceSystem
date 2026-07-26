"""
Assorting Entry module — endpoints.
Uses the project standard: mysql.connector via src.database.get_db.

Routes (mounted under /assorting):
    GET    /assorting/sheds              DISTINCT shed_type from machine_mst by branch
    GET    /assorting/machines           Mc Nos (machine_mst, machine_type_id=8,
                                                 filtered by branch + optional shed_type)
    GET    /assorting/qualities          Jute qualities (jute_quality_mst)
    GET    /assorting/selectors          Selectors (tbl_selector_mst)
    GET    /assorting/validate-trolly    Validate trolly (trolly_mst, trolly_type='S')
    POST   /assorting/entry              Save an assorting entry row
    GET    /assorting/entry              List entries for date+branch (+optional mc)
    DELETE /assorting/entry/<id>         Delete a row
"""
from flask import request, jsonify
from . import assorting_bp
from src.database import get_db


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _ensure_assorting_table():
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS assorting_entry (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                entry_date      DATE NOT NULL,
                shed_type       VARCHAR(32),
                branch_id       INT,
                mc_id           INT,
                selector_id     INT,
                quality_id      INT,
                trolly_id       INT,
                trolly_no       VARCHAR(64),
                gross_wt        DECIMAL(12,3),
                tare_wt         DECIMAL(12,3),
                net_wt          DECIMAL(12,3),
                user_id         INT,
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_assorting_date_branch (entry_date, branch_id)
            )
        """)
        db.commit()
        cursor.close()
        db.close()
    except Exception as e:
        print(f"⚠️  assorting_entry create error (non-fatal): {e}")


# ─── Sheds ────────────────────────────────────────────────────────────────────

@assorting_bp.route('/sheds', methods=['GET'])
def get_assorting_sheds():
    """
    Distinct shed types for Assorting Entry.
    Source: machine_mst — DISTINCT shed_type, filtered by branch_id AND machine_type_id = 8.
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        if not branch_id:
            return jsonify({'status': 'error', 'message': 'branch_id is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        sql="""
            SELECT DISTINCT shed_type
            FROM machine_mst mm
            left join dept_mst dm on mm.dept_id =dm.dept_id  
            WHERE dm.branch_id =%s
              AND machine_type_id = 8
              AND shed_type IS NOT NULL
              AND TRIM(shed_type) <> ''
              AND (active IS NULL OR active = 1)
            ORDER BY shed_type
        """
        #print(f"Executing SQL: {sql} with params: {(branch_id,)}")
        cursor.execute(sql, (branch_id,))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({
            'status': 'success',
            'sheds': [r['shed_type'] for r in rows],
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Machines (Mc Nos) ────────────────────────────────────────────────────────

@assorting_bp.route('/machines', methods=['GET'])
def get_assorting_machines():
    """
    Mc Nos for Assorting Entry.
    Source: machine_mst where machine_type_id = 8 and branch_id (and shed_type if given).
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        shed_type = (request.args.get('shed_type') or '').strip() or None
        if not branch_id:
            return jsonify({'status': 'error', 'message': 'branch_id is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        sql="""
            SELECT mm.machine_id     AS mc_id,
                   mm.machine_name   AS mc_name,
                   mm.mech_code      AS mc_code,
                   mm.mech_shr_code  AS mc_short,
                   mm.shed_type      AS shed_type,
                   mm.dept_id        AS dept_id,
                   dm.dept_desc      AS dept_name
            FROM machine_mst mm
            LEFT JOIN dept_mst dm ON mm.dept_id = dm.dept_id
            WHERE mm.machine_type_id in(8,50)
              AND dm.branch_id = %s
              AND (%s IS NULL OR mm.shed_type = %s)
              AND (mm.active IS NULL OR mm.active = 1)
            ORDER BY mm.mech_code, mm.machine_name
        """
        #print(f"Executing SQL: {sql} with params: {(branch_id, shed_type, shed_type)}")
        cursor.execute(sql, (branch_id, shed_type, shed_type))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'machines': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Jute Qualities ───────────────────────────────────────────────────────────

@assorting_bp.route('/qualities', methods=['GET'])
def get_assorting_qualities():
    """
    Jute qualities for Assorting Entry.
    Source: jute_quality_mst — shr_name shown as button label, filtered by branch.
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        if not branch_id:
            return jsonify({'status': 'error', 'message': 'branch_id is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT jute_qlty_id AS quality_id,
                   COALESCE(shr_name, jute_quality) AS shr_name,
                   jute_quality AS quality_name
            FROM jute_quality_mst
            WHERE branch_id = %s
              AND active = 1
            ORDER BY shr_name, jute_quality
        """, (branch_id,))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'qualities': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Selectors ────────────────────────────────────────────────────────────────

@assorting_bp.route('/selectors', methods=['GET'])
def get_assorting_selectors():
    """
    Selectors for Assorting Entry.
    Source: tbl_selector_mst — selector_shr_name shown as button label, filtered by branch.
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        if not branch_id:
            return jsonify({'status': 'error', 'message': 'branch_id is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        sql="""SELECT tbl_selector_mst_id selector_id,
                   COALESCE(selector_shr_name, selector_name) AS selector_shr_name,
                   selector_name
            FROM tbl_selector_mst
            WHERE branch_id = %s
              AND (active IS NULL OR active = 1)
            ORDER BY selector_shr_name, selector_name
            """
        cursor.execute(sql, (branch_id,))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'selectors': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Validate Trolly (trolly_type='S') ────────────────────────────────────────

@assorting_bp.route('/validate-trolly', methods=['GET'])
def validate_assorting_trolly():
    """
    Validate a trolly number against trolly_mst for trolly_type='S'.
    Returns trolly_id and tare weight if found.
    """
    try:
        trolly_no = (request.args.get('trolly_no') or '').strip()
        branch_id = request.args.get('branch_id', type=int)
        if not trolly_no:
            return jsonify({'status': 'error', 'message': 'trolly_no is required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        sql="""
            SELECT trolly_id,
                   ' ' trolly_no,
                   trolly_name,
                   trolly_weight,
                   trolly_type,
                   branch_id
            FROM trolly_mst
            WHERE TRIM(trolly_posting_code) = %s
              AND trolly_type = 'A'
              AND (%s IS NULL OR branch_id = %s)
       --       AND (active IS NULL OR active = 1)
            LIMIT 1
        """
        #print(f"Executing SQL: {sql} with params: {(trolly_no, branch_id, branch_id)}")
        cursor.execute(sql, (trolly_no, branch_id, branch_id))
        row = cursor.fetchone()
        cursor.close()
        db.close()

        if not row:
            return jsonify({'status': 'error', 'message': 'Trolly not found'}), 200

        return jsonify({
            'status': 'success',
            'trolly_id':     row.get('trolly_id'),
            'trolly_no':     row.get('trolly_no'),
            'trolly_name':   row.get('trolly_name'),
            'trolly_weight': row.get('trolly_weight'),
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Save Assorting Entry ─────────────────────────────────────────────────────

@assorting_bp.route('/entry', methods=['POST'])
def save_assorting_entry():
    """
    Save an Assorting Entry row.
    Body JSON: { date, shed_type, branch_id, mc_id, selector_id, quality_id,
                 trolly_id, trolly_no, gross_wt, tare_wt, net_wt, user_id }
    """
    try:
        _ensure_assorting_table()
        data = request.get_json() or {}
        date_str    = data.get('date')
        shed_type   = (data.get('shed_type') or '').strip() or None
        branch_id   = data.get('branch_id')
        mc_id       = data.get('mc_id')
        selector_id = data.get('selector_id')
        quality_id  = data.get('quality_id')
        trolly_id   = data.get('trolly_id')
        trolly_no   = (data.get('trolly_no') or '').strip() or None
        gross_wt    = data.get('gross_wt')
        tare_wt     = data.get('tare_wt')
        net_wt      = data.get('net_wt')
        user_id     = data.get('user_id')

        if not all([date_str, branch_id, mc_id, selector_id, quality_id]):
            return jsonify({'status': 'error',
                            'message': 'date, branch_id, mc_id, selector_id, quality_id are required'}), 400

        try:
            gross_wt = float(gross_wt) if gross_wt is not None else 0.0
            tare_wt  = float(tare_wt)  if tare_wt  is not None else 0.0
            net_wt   = float(net_wt)   if net_wt   is not None else (gross_wt - tare_wt)
        except Exception:
            return jsonify({'status': 'error', 'message': 'weights must be numeric'}), 400

        if net_wt < 0:
            return jsonify({'status': 'error', 'message': 'Net weight cannot be negative'}), 400

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO assorting_entry
                (entry_date, shed_type, branch_id, mc_id, selector_id, quality_id,
                 trolly_id, trolly_no, gross_wt, tare_wt, net_wt, user_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (date_str, shed_type, branch_id, mc_id, selector_id, quality_id,
              trolly_id, trolly_no, gross_wt, tare_wt, net_wt, user_id))
        db.commit()
        new_id = cursor.lastrowid
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'message': 'Saved', 'id': new_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── List Assorting Entries ───────────────────────────────────────────────────

@assorting_bp.route('/entry', methods=['GET'])
def list_assorting_entries():
    """
    List assorting entries for a date and branch (optionally filtered by mc_id).
    """
    try:
        _ensure_assorting_table()
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        mc_id     = request.args.get('mc_id', type=int)
        if not all([date_str, branch_id]):
            return jsonify({'status': 'error',
                            'message': 'date and branch_id are required'}), 400

        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT a.id,
                   a.entry_date,
                   a.shed_type,
                   a.mc_id,
                   COALESCE(mm.mech_code, mm.machine_name, CONCAT('MC', a.mc_id)) AS mc_no,
                   a.selector_id,
                   COALESCE(sm.selector_shr_name, sm.selector_name, '')           AS selector_name,
                   a.quality_id,
                   COALESCE(jq.shr_name, jq.jute_quality, '')                     AS quality,
                   a.trolly_id,
                   a.trolly_no,
                   a.gross_wt,
                   a.tare_wt,
                   a.net_wt
            FROM assorting_entry a
            LEFT JOIN machine_mst       mm ON a.mc_id       = mm.machine_id
            LEFT JOIN tbl_selector_mst  sm ON a.selector_id = sm.tbl_selector_mst_id
            LEFT JOIN jute_quality_mst  jq ON a.quality_id  = jq.jute_qlty_id
            WHERE a.entry_date = %s
              AND a.branch_id  = %s
              AND (%s IS NULL OR a.mc_id = %s)
            ORDER BY a.id DESC
        """, (date_str, branch_id, mc_id, mc_id))
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'entries': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Assorting Report (production by selector / machine-type / quality) ────────

def _assorting_report_sections(date_str, branch_id):
    """Return (by_selector, by_machine_type, by_quality) production sections,
    each {rows: [{name, production}], grand_total}. Shared by the JSON and
    .xlsx endpoints so they can never drift."""
    db = get_db()
    cur = db.cursor(dictionary=True)

    def rows(sql):
        cur.execute(sql, (date_str, branch_id))
        out = [{'name': (r['name'] or ''), 'production': float(r['production'] or 0)}
               for r in cur.fetchall()]
        return {'rows': out, 'grand_total': round(sum(x['production'] for x in out), 3)}

    try:
        # 1) Selector wise
        by_selector = rows("""
            SELECT COALESCE(tsm.selector_name, '') AS name, SUM(ae.net_wt) AS production
            FROM assorting_entry ae
            LEFT JOIN tbl_selector_mst tsm ON ae.selector_id = tsm.tbl_selector_mst_id
            WHERE ae.entry_date = %s AND ae.branch_id = %s
            GROUP BY tsm.selector_name
            ORDER BY production DESC
        """)
        # 2) Machine type wise (machine_type_id 8 = Spreader, else Softner)
        by_machine_type = rows("""
            SELECT CASE WHEN g.machine_type_id = 8 THEN 'Spreader' ELSE 'Softner' END AS name,
                   g.production
            FROM (
                SELECT mm.machine_type_id, SUM(ae.net_wt) AS production
                FROM assorting_entry ae
                LEFT JOIN machine_mst mm ON ae.mc_id = mm.machine_id
                WHERE ae.entry_date = %s AND ae.branch_id = %s
                GROUP BY mm.machine_type_id
            ) g
            ORDER BY g.production DESC
        """)
        # 3) Quality wise
        by_quality = rows("""
            SELECT COALESCE(jqm.shr_name, '') AS name, SUM(ae.net_wt) AS production
            FROM assorting_entry ae
            LEFT JOIN jute_quality_mst jqm ON ae.quality_id = jqm.jute_qlty_id
            WHERE ae.entry_date = %s AND ae.branch_id = %s
            GROUP BY jqm.shr_name
            ORDER BY production DESC
        """)
        return by_selector, by_machine_type, by_quality
    finally:
        cur.close(); db.close()


def _build_assorting_xlsx(sections, date_str):
    """Build the Assorting report workbook: three Name|Production blocks
    (Selector / Machine Type / Quality), each with a Total -> BytesIO."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def z(v):
        v = round(float(v or 0), 3)
        return None if v == 0 else (int(v) if v == int(v) else v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assorting"
    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1565C0')
    hdr_font = Font(bold=True, color='FFFFFF')
    sec_fill = PatternFill('solid', fgColor='BBDEFB')
    tot_fill = PatternFill('solid', fgColor='E3F2FD')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.cell(1, 1, f"Assorting Report - Date: {date_str}").font = Font(bold=True)
    ws.merge_cells('A1:B1')
    ws.cell(1, 1).alignment = center

    r = 3
    for title, sec in zip(('Selector wise', 'Machine Type wise', 'Quality wise'), sections):
        for c in (1, 2):
            ws.cell(r, c).fill = sec_fill; ws.cell(r, c).border = border
        sh = ws.cell(r, 1, title); sh.font = Font(bold=True); sh.alignment = left
        r += 1
        for c, h in enumerate(('Particular', 'Production'), start=1):
            cell = ws.cell(r, c, h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border
        r += 1
        for row in sec['rows']:
            a = ws.cell(r, 1, row['name']); a.alignment = left; a.border = border
            b = ws.cell(r, 2, z(row['production'])); b.alignment = center; b.border = border
            r += 1
        ta = ws.cell(r, 1, 'Total'); ta.font = Font(bold=True); ta.fill = tot_fill
        ta.alignment = left; ta.border = border
        tb = ws.cell(r, 2, z(sec['grand_total'])); tb.font = Font(bold=True)
        tb.fill = tot_fill; tb.alignment = center; tb.border = border
        r += 2   # blank gap between sections

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@assorting_bp.route('/report', methods=['GET'])
def assorting_report():
    """Three production reports for a date+branch, each with a grand total:
       1) by selector, 2) by machine type, 3) by jute quality.
    """
    try:
        _ensure_assorting_table()
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not all([date_str, branch_id]):
            return jsonify({'status': 'error',
                            'message': 'date and branch_id are required'}), 400

        by_selector, by_machine_type, by_quality = _assorting_report_sections(date_str, branch_id)
        return jsonify({
            'status':          'success',
            'by_selector':     by_selector,
            'by_machine_type': by_machine_type,
            'by_quality':      by_quality,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@assorting_bp.route('/report.xlsx', methods=['GET'])
def assorting_report_xlsx():
    """Assorting report as a downloadable .xlsx (3 sections in one sheet)."""
    try:
        from flask import send_file
        _ensure_assorting_table()
        date_str  = request.args.get('date')
        branch_id = request.args.get('branch_id', type=int)
        if not all([date_str, branch_id]):
            return jsonify({'status': 'error',
                            'message': 'date and branch_id are required'}), 400

        sections = _assorting_report_sections(date_str, branch_id)
        buf = _build_assorting_xlsx(sections, date_str)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"AssortingReport_{date_str}.xlsx",
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Update Assorting Entry ───────────────────────────────────────────────────

@assorting_bp.route('/entry/<int:entry_id>', methods=['PUT'])
def update_assorting_entry(entry_id):
    """
    Update an existing Assorting Entry row.
    Body JSON: { date, shed_type, branch_id, mc_id, selector_id, quality_id,
                 trolly_id, trolly_no, gross_wt, tare_wt, net_wt, user_id }
    """
    try:
        data = request.get_json() or {}
        date_str    = data.get('date')
        shed_type   = (data.get('shed_type') or '').strip() or None
        branch_id   = data.get('branch_id')
        mc_id       = data.get('mc_id')
        selector_id = data.get('selector_id')
        quality_id  = data.get('quality_id')
        trolly_id   = data.get('trolly_id')
        trolly_no   = (data.get('trolly_no') or '').strip() or None
        gross_wt    = data.get('gross_wt')
        tare_wt     = data.get('tare_wt')
        net_wt      = data.get('net_wt')
        user_id     = data.get('user_id')

        if not all([date_str, branch_id, mc_id, selector_id, quality_id]):
            return jsonify({'status': 'error',
                            'message': 'date, branch_id, mc_id, selector_id, quality_id are required'}), 400

        try:
            gross_wt = float(gross_wt) if gross_wt is not None else 0.0
            tare_wt  = float(tare_wt)  if tare_wt  is not None else 0.0
            net_wt   = float(net_wt)   if net_wt   is not None else (gross_wt - tare_wt)
        except Exception:
            return jsonify({'status': 'error', 'message': 'weights must be numeric'}), 400

        if net_wt < 0:
            return jsonify({'status': 'error', 'message': 'Net weight cannot be negative'}), 400

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            UPDATE assorting_entry
               SET entry_date  = %s,
                   shed_type   = %s,
                   branch_id   = %s,
                   mc_id       = %s,
                   selector_id = %s,
                   quality_id  = %s,
                   trolly_id   = %s,
                   trolly_no   = %s,
                   gross_wt    = %s,
                   tare_wt     = %s,
                   net_wt      = %s,
                   user_id     = %s
             WHERE id = %s
        """, (date_str, shed_type, branch_id, mc_id, selector_id, quality_id,
              trolly_id, trolly_no, gross_wt, tare_wt, net_wt, user_id, entry_id))
        affected = cursor.rowcount
        db.commit()
        cursor.close()
        db.close()

        if affected == 0:
            return jsonify({'status': 'error', 'message': 'Entry not found'}), 404
        return jsonify({'status': 'success', 'message': 'Updated', 'id': entry_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─── Delete Assorting Entry ───────────────────────────────────────────────────

@assorting_bp.route('/entry/<int:entry_id>', methods=['DELETE'])
def delete_assorting_entry(entry_id):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("DELETE FROM assorting_entry WHERE id = %s", (entry_id,))
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'status': 'success', 'message': 'Deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
