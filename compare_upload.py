"""
Upload-and-compare: parse a DailyAttendance_DetailedReport .xls, apply the
shift conversion rule (shift A with OT > 0 -> B1, working_hours = Work Dur.),
compare against the daily_attendance table (13.126.47.172 / sjm), and write
a downloadable comparison Excel.

Usage:
    python compare_upload.py "DailyAttendance_DetailedReport (7).xls" 2026-07-01
    python compare_upload.py "report.xls"            # all dates in the file

Output: Attendance_Compare_<date>.xlsx in the current folder.

Requires: pip install pandas xlrd openpyxl mysql-connector-python
"""
import re
import sys

import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DB = dict(host='13.126.47.172', user='myroot', password='deb#9876',
          database='sjm')


def hm_to_hours(v):
    """'8:35' -> 8.58 ; invalid/blank -> None"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    m = re.match(r'^(\d+):(\d{2})$', str(v).strip())
    return round(int(m.group(1)) + int(m.group(2)) / 60, 2) if m else None


def parse_report(path, only_date=None):
    """Parse the multi-section detailed report. Returns list of dicts."""
    df = pd.read_excel(path, header=None)
    rows, date, dept = [], None, None
    for i in range(len(df)):
        c1 = df.iat[i, 1]
        if isinstance(c1, str) and 'Attendance Date' in c1:
            date = pd.to_datetime(str(df.iat[i, 5]).strip(),
                                  format='%d-%b-%Y').date()
            continue
        if isinstance(c1, str) and c1.strip() == 'Department':
            dept = str(df.iat[i, 3]).strip()
            continue
        if isinstance(c1, str) and c1.strip() == 'SNo':
            continue
        try:
            int(c1)
        except (TypeError, ValueError):
            continue
        ec = df.iat[i, 2]
        if pd.isna(ec) or (only_date and date != only_date):
            continue
        shift = str(df.iat[i, 5]).strip() if pd.notna(df.iat[i, 5]) else ''
        work = str(df.iat[i, 12]).strip() if pd.notna(df.iat[i, 12]) else ''
        ot = str(df.iat[i, 13]).strip() if pd.notna(df.iat[i, 13]) else ''
        wh, oth = hm_to_hours(work), hm_to_hours(ot) or 0
        spell = 'B1' if shift == 'A' and oth > 0 else shift
        rows.append(dict(
            date=date, dept=dept, ecode=str(ec).strip(),
            name=str(df.iat[i, 3]).strip(), shift=shift, work=work, ot=ot,
            work_hrs=wh, ot_hrs=oth, spell=spell, working_hours=wh,
            status=str(df.iat[i, 17]).strip() if pd.notna(df.iat[i, 17]) else ''))
    return rows


def fetch_db(dates):
    """{(ecode_no_leading_zeros, 'YYYY-MM-DD'): (spell, hours, name)}"""
    cnx = mysql.connector.connect(**DB)
    cur = cnx.cursor()
    cur.execute("""
        SELECT p.emp_code, da.attendance_date, da.spell,
               COALESCE(da.working_hours, 0),
               CONCAT(p.first_name, ' ', COALESCE(p.last_name, ''))
        FROM daily_attendance da
        JOIN hrms_ed_personal_details p ON da.eb_id = p.eb_id
        WHERE da.attendance_date BETWEEN %s AND %s AND da.is_active = 1
    """, (min(dates), max(dates)))
    out = {}
    for ec, d, spell, wh, nm in cur.fetchall():
        if d in dates:
            out[(str(ec).lstrip('0'), str(d))] = (spell or '', float(wh), nm)
    cnx.close()
    return out


HDR = ['attendance_date', 'emp_code', 'emp_name', 'department',
       'report_shift', 'report_ot', 'spell (converted)',
       'working_hours (report)', 'spell (DB)', 'working_hours (DB)',
       'spell_match', 'hours_match', 'result']
RED = PatternFill('solid', start_color='FFC7CE')
ORANGE = PatternFill('solid', start_color='FFEB9C')
BLUE = PatternFill('solid', start_color='DDEBF7')
THIN = Border(*[Side(style='thin')] * 4)


def write_xlsx(rows, dbmap, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparison'
    ws.append(HDR)
    for c in ws[1]:
        c.font = Font(name='Arial', bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='305496')
        c.alignment = Alignment(horizontal='center')
        c.border = THIN
    ok = mism = miss = 0
    seen = set()
    for r in rows:
        key = (r['ecode'].lstrip('0'), str(r['date']))
        seen.add(key)
        hit = dbmap.get(key)
        if hit is None:
            ws.append([r['date'], r['ecode'], r['name'], r['dept'],
                       r['shift'], r['ot'], r['spell'], r['working_hours'],
                       'NOT IN DB', '', '', '', 'MISSING IN DB'])
            fill, miss = ORANGE, miss + 1
        else:
            dsp, dwh, _ = hit
            sm = 'Y' if dsp.strip().upper() == r['spell'].upper() else 'N'
            hm = 'Y' if abs((r['working_hours'] or 0) - dwh) < 0.02 else 'N'
            res = 'OK' if sm == hm == 'Y' else 'MISMATCH'
            ws.append([r['date'], r['ecode'], r['name'], r['dept'],
                       r['shift'], r['ot'], r['spell'], r['working_hours'],
                       dsp, dwh, sm, hm, res])
            if res == 'OK':
                fill, ok = None, ok + 1
            else:
                fill, mism = RED, mism + 1
        if fill:
            for c in ws[ws.max_row]:
                c.fill = fill
    extra = 0
    for key, (dsp, dwh, nm) in sorted(dbmap.items()):
        if key not in seen:
            ws.append([key[1], key[0], nm, '', '', '', '', '',
                       dsp, dwh, '', '', 'DB ONLY'])
            for c in ws[ws.max_row]:
                c.fill = BLUE
            extra += 1
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = THIN
            c.font = Font(name='Arial')
        row[0].number_format = 'DD-MMM-YYYY'
    for col, w in zip('ABCDEFGHIJKLM',
                      [14, 10, 28, 18, 12, 10, 15, 18, 12, 16, 11, 11, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:M{ws.max_row}'

    s = wb.create_sheet('Summary')
    for row in [('Rows in report', len(rows)), ('OK (spell + hours match)', ok),
                ('MISMATCH', mism), ('MISSING IN DB', miss),
                ('DB ONLY (not in report)', extra)]:
        s.append(row)
    for c in s['A']:
        c.font = Font(name='Arial', bold=True)
    s.column_dimensions['A'].width = 28
    wb.save(out_path)
    return ok, mism, miss, extra


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    path = sys.argv[1]
    only = (pd.to_datetime(sys.argv[2]).date() if len(sys.argv) > 2 else None)
    rows = parse_report(path, only)
    if not rows:
        sys.exit(f'No rows found for {only} in {path}')
    dates = sorted({r['date'] for r in rows})
    print(f'Parsed {len(rows)} rows for {", ".join(map(str, dates))} '
          f'(A->B1 conversions: {sum(1 for r in rows if r["spell"] == "B1")})')
    dbmap = fetch_db(set(dates))
    print(f'DB rows fetched: {len(dbmap)}')
    out = f'Attendance_Compare_{only or "all"}.xlsx'
    ok, mism, miss, extra = write_xlsx(rows, dbmap, out)
    print(f'OK: {ok}  MISMATCH: {mism}  MISSING IN DB: {miss}  '
          f'DB ONLY: {extra}\nSaved: {out}')


if __name__ == '__main__':
    main()
