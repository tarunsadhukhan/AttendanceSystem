"""
Compare converted report (DailyAttendance_Converted.xlsx) against the
daily_attendance table on 13.126.47.172 / sjm.
Run locally:  python compare_daily_attendance.py
Adds a 'DB_Compare' sheet: report vs DB spell + working_hours per emp/date.
"""
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = 'DailyAttendance_Converted.xlsx'

db = mysql.connector.connect(host='13.126.47.172', user='myroot',
                             password='deb#9876', database='sjm')
cur = db.cursor()
cur.execute("""
    SELECT p.emp_code, da.attendance_date, da.spell,
           COALESCE(da.working_hours, 0)
    FROM daily_attendance da
    JOIN hrms_ed_personal_details p ON da.eb_id = p.eb_id
    WHERE da.attendance_date BETWEEN '2026-07-01' AND '2026-07-05'
      AND da.is_active = 1
""")
dbmap = {}
for ec, d, spell, wh in cur.fetchall():
    dbmap[(str(ec).lstrip('0'), str(d))] = (spell or '', float(wh))
print(f"DB rows fetched: {len(dbmap)}")

wb = load_workbook(XLSX, data_only=True)
src = wb['daily_attendance']
if 'DB_Compare' in wb.sheetnames:
    del wb['DB_Compare']
ws = wb.create_sheet('DB_Compare')
ws.append(['attendance_date', 'emp_code', 'emp_name', 'report_spell',
           'db_spell', 'report_hours', 'db_hours', 'spell_match',
           'hours_match', 'result'])
thin = Border(*[Side(style='thin')] * 4)
for c in ws[1]:
    c.font = Font(name='Arial', bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='305496')
    c.alignment = Alignment(horizontal='center')
    c.border = thin
red = PatternFill('solid', start_color='FFC7CE')
orange = PatternFill('solid', start_color='FFEB9C')
mism = miss = ok = 0
for r in src.iter_rows(min_row=2, values_only=True):
    date, ec, name, spell, wh = r[0], str(r[1]), r[2], r[10], r[11]
    dkey = (ec.lstrip('0'), str(date.date() if hasattr(date, 'date') else date))
    hit = dbmap.get(dkey)
    if hit is None:
        ws.append([date, ec, name, spell, 'NOT IN DB', wh, '', '', '', 'MISSING'])
        for c in ws[ws.max_row]:
            c.fill = orange
        miss += 1
    else:
        dsp, dwh = hit
        sm = 'Y' if str(dsp).strip().upper() == str(spell).strip().upper() else 'N'
        hm = 'Y' if abs((wh or 0) - dwh) < 0.02 else 'N'
        res = 'OK' if sm == 'Y' and hm == 'Y' else 'MISMATCH'
        ws.append([date, ec, name, spell, dsp, wh, dwh, sm, hm, res])
        if res == 'MISMATCH':
            for c in ws[ws.max_row]:
                c.fill = red
            mism += 1
        else:
            ok += 1
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.border = thin
        if c.column == 1:
            c.number_format = 'DD-MMM-YYYY'
for col, w in zip('ABCDEFGHIJ', [14, 10, 28, 12, 12, 13, 10, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:J{ws.max_row}'
wb.save(XLSX)
print(f"OK: {ok}  MISMATCH: {mism}  NOT IN DB: {miss}")
print(f"Comparison written to '{XLSX}' -> sheet 'DB_Compare'")
